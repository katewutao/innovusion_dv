import openpyxl
import re
from openpyxl.styles import PatternFill, Side, Border, Alignment
from openpyxl.utils import get_column_letter
import platform
import datetime


class ExcelFormat(object):
    def __init__(self, excel_path, save_path=False):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(self.excel_path)
        if not save_path:
            self.save_path = excel_path
        else:
            self.save_path = save_path

    def lens(self, str1):
        str1 = str(str1)
        l = 0
        for i in range(len(str1)):
            if ord(str1[i]) < 256:
                l += 1
            else:
                l += 2
        return l

    def value_float(self):
        border_style = Side(style='thin', color='000000')
        border = Border(left=border_style, right=border_style,
                        top=border_style, bottom=border_style)
        align = Alignment(horizontal="left")
        for sheetname in self.wb.sheetnames:
            sheet = self.wb[sheetname]
            for i in range(sheet.max_row):
                for j in range(sheet.max_column):
                    cell = sheet.cell(row=i+1, column=j+1)
                    cell.border = border
                    cell.alignment = align
                    if re.search("^-?\d+\.?\d*$", str(cell.value)):
                        cell.value = float(cell.value)

    def column_width(self):
        if "windows" in platform.platform().lower():
            compensate = 1.4
        else:
            compensate = 1.3
        for sheetname in self.wb.sheetnames:
            sheet = self.wb[sheetname]
            column_max_width = [0]*sheet.max_column
            for i in range(sheet.max_row):
                for j in range(sheet.max_column):
                    cell = sheet.cell(row=i+1, column=j+1)
                    current_width = self.lens(cell.value)
                    if current_width < 50:
                        column_max_width[j] = max(
                            column_max_width[j], current_width)
            for j in range(sheet.max_column):
                sheet.column_dimensions[get_column_letter(
                    j+1)].width = column_max_width[j]*compensate

    def fill_background_color(self):
        fill_red = PatternFill(start_color='FF0000',
                               end_color='FF0000', fill_type='darkGray')
        fill_yellow = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='darkGray')
        fill_green = PatternFill(start_color='00FF00',
                                 end_color='00FF00', fill_type='darkGray')
        fill_white = PatternFill(start_color='FFFFFF',
                                 end_color='FFFFFF', fill_type='darkGray')
        for sheetname in self.wb.sheetnames:
            sheet = self.wb[sheetname]
            # first column and first row check
            for i in range(sheet.max_row):
                cell = sheet.cell(row=i+1, column=1)
                if type(cell.value) != type(None) and (type(cell.value) != str or cell.value.strip() != ""):
                    cell.fill = fill_green
                else:
                    cell.fill = fill_white
            for j in range(sheet.max_column):
                cell = sheet.cell(row=1, column=j+1)
                if type(cell.value) != type(None) and (type(cell.value) != str or cell.value.strip() != ""):
                    cell.fill = fill_green
                else:
                    cell.fill = fill_white
            # value check
            for i in range(1, sheet.max_row):
                ll = sheet.cell(row=i+1, column=2).value
                ul = sheet.cell(row=i+1, column=4).value
                current = sheet.cell(row=i+1, column=7)
                result_type = [float, int]
                pass_flag = False
                if type(ll) not in result_type and type(ul) not in result_type:
                    current.fill = fill_white
                    continue
                if type(current.value) in result_type:
                    pass_flag = True
                    if (type(ll) in result_type and current.value < ll) or (type(ul) in result_type and current.value > ul):
                        pass_flag = False
                if pass_flag:
                    current.fill = fill_green
                else:
                    current.fill = fill_yellow
            # pass check
            last_cell = sheet.cell(row=sheet.max_row, column=sheet.max_column)
            if type(last_cell.value) == str:
                if last_cell.value.lower() == "fail":
                    last_cell.fill = fill_yellow
                elif last_cell.value.lower() == "pass":
                    last_cell.fill = fill_green

    def format(self):
        self.column_width()
        self.value_float()
        self.wb.save(self.save_path)